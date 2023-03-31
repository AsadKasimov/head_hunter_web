import os
import re
from PyPDF2 import PdfFileReader
import csv
from openpyxl import load_workbook

path_to_file_ex = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'hh_ru', 'data',
                               'test_with_files.xlsx')
path_to_file_pdf = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'hh_ru', 'data',
                                'simple_demo.pdf')
path_to_file_csv = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'hh_ru', 'data',
                                'some.csv')


def test_csv():
    with open(path_to_file_csv, newline='') as f:
        for row in csv.reader(f):
            assert 'California' in row
            assert ['California', 'Sacramento', 'Los Angeles', '39538223'] == row
            assert len(row) == 4


# тут работа с pdf
def test_pdf():
    with open(path_to_file_pdf, "rb") as filehandle:
        pdf = PdfFileReader(filehandle)
        page1 = pdf.getPage(0)
        text = page1.extractText()

        print(re.findall(r'^\w+', text, flags=re.IGNORECASE) == ['Welcome'])
        text_from = re.findall(r'\w+', text, flags=re.IGNORECASE)
        assert 'Welcome' in text_from
        assert 'to' in text_from
        assert 'Python' in text_from
        assert len(text_from) == 3


def test_ex():
    book = load_workbook(path_to_file_ex)
    sheet = book.active
    ass = sheet.cell(row=3, column=3).value
    assert sheet.max_row == 11
    assert sheet.max_column == 8
    assert ass == 25000
